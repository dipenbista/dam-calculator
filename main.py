"""
main.py — FastAPI backend for the Gravity Dam Stability Calculator
Run with:  uvicorn main:app --reload
"""

from fastapi import FastAPI, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field
from typing import List, Tuple, Optional
import traceback
import io
import base64
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, numbers)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

from calculator import (
    DamGeometry, MaterialProperties, WaterLevels,
    DrainageConfig, SiltConfig, BackfillConfig,
    IcePressureConfig, RockBoltConfig, RockAnchorConfig,
    AppliedForceConfig, EarthquakeConfig, run_load_case, plot_to_base64,
    generate_detailed_calc, polygon_area_centroid
)

app = FastAPI(title="Gravity Dam Stability Calculator")

# Serve the frontend (index.html and any other static files)
app.mount("/static", StaticFiles(directory="static"), name="static")


@app.get("/")
def root():
    return FileResponse("static/index.html")


# =============================================================================
# REQUEST / RESPONSE MODELS
# =============================================================================

class CoordPoint(BaseModel):
    x: float
    y: float

class LoadCaseRequest(BaseModel):
    model_config = {"extra": "ignore"}   # silently drop unknown fields → no 422 on version skew
    # Geometry
    coordinates:         List[CoordPoint]
    upstream_top_point:  CoordPoint

    # Water levels
    HRV_us: float;  DFV_us: float;  MFV_us: float
    HRV_ds: float = 0.0
    DFV_ds: float = 0.0
    MFV_ds: float = 0.0

    # Material
    unit_weight_dam:   float = 24.0
    unit_weight_water: float = 10.0
    friction_coeff:    float = 0.70

    # Drainage
    drainage_include:            bool  = False
    drainage_distance_from_heel: float = 2.0
    drainage_reduction_factor:   float = 0.333

    # Silt
    silt_include:               bool  = False
    silt_height_us:             float = 0.0     # height above heel (legacy)
    silt_elevation_us:          float = 0.0     # absolute elevation of silt surface (preferred)
    silt_unit_weight_submerged: float = 9.0
    silt_phi_deg:               float = 30.0

    # Backfill
    backfill_include_pressure:      bool  = False
    backfill_include_weight:        bool  = False
    backfill_height:                float = 0.0     # height above toe (legacy)
    backfill_elevation:             float = 0.0     # absolute elevation of backfill surface (preferred)
    backfill_coeff_pressure:        float = 0.333
    backfill_unit_weight_dry:       float = 18.0
    backfill_unit_weight_wet:       float = 20.0
    backfill_unit_weight_submerged: float = 10.0

    # Ice
    ice_include:   bool  = False
    ice_pressure:  float = 150.0

    # Rock bolt
    rock_bolt_include:              bool  = False
    rock_bolt_force_per_m:          float = 0.0
    rock_bolt_cover_from_heel:      float = 0.5
    rock_bolt_apply_depth_limit:    bool  = True   # apply regulatory depth limit
    rock_bolt_depth_limit:          float = 7.0    # limit in m (default NVE 7 m)

    # Rock anchor
    rock_anchor_include:         bool  = False
    rock_anchor_force_per_m:     float = 0.0
    rock_anchor_cover_from_heel: float = 2.0

    # Applied external forces
    # vertical_forces   : list of [force_kN_per_m, distance_from_toe_m]
    # horizontal_forces : list of [force_kN_per_m, height_above_toe_m]
    #   positive V = downward (stabilising), negative V = upward (destabilising)
    #   positive H = acts toward upstream (destabilising), negative = toward DS (stabilising)
    applied_vertical_forces:   List[List[float]] = Field(default_factory=list)
    applied_horizontal_forces: List[List[float]] = Field(default_factory=list)

    # Which load cases to run
    run_HRV:             bool = True
    run_DFV:             bool = True
    run_MFV:             bool = True
    run_DFV_no_bolts:    bool = False
    run_EQ:              bool = False
    eq_a_h:              float = 0.0
    eq_a_v:              float = 0.0
    fs_uls:              float = 1.5
    fs_als:              float = 1.1
    res_uls:             str   = 'middle_third'
    res_als:             str   = 'l6'


class ForceRow(BaseModel):
    name:         str
    V:            float
    H:            float
    x_from_toe:   float
    y_from_toe:   float
    M_res:        float
    M_ov:         float
    stabilising:  bool

class CaseResult(BaseModel):
    case_name:            str
    sum_V:                float
    H_net:                float
    sum_M_res:            float
    sum_M_ov:             float
    x_resultant:          float
    eccentricity:         float
    in_middle_third:      bool
    resultant_check_type: str
    sigma_toe:            float
    sigma_heel:           float
    FS_sliding:           float
    FS_overturning:       float
    tension_length:       float
    forces:               List[ForceRow]
    messages:             List[dict]
    fs_threshold:         float = 1.5   # engineering messages for this case
    plot_dam_base64:      str = ''     # dam cross-section PNG (base64)
    plot_stress_base64:   str = ''     # base stress PNG (base64)

class GeometryInfo(BaseModel):
    toe_elevation:          float
    heel_elevation:         float
    base_length_horizontal: float
    dam_height:             float
    warnings:               List[str]   # kept for backward compat (now empty)

class CalculationResponse(BaseModel):
    geometry: GeometryInfo
    results:  List[CaseResult]


# =============================================================================
# MAIN ENDPOINT
# =============================================================================

def _resolve_silt_height(req, geom) -> float:
    """Silt height above heel; prefer absolute elevation, fall back to legacy height."""
    if not req.silt_include:
        return 0.0
    if getattr(req, "silt_elevation_us", 0) and req.silt_elevation_us > 0:
        h = req.silt_elevation_us - geom.heel_elevation
    else:
        h = req.silt_height_us
    return max(h, 0.0)


def _resolve_backfill_height(req, geom) -> float:
    """Backfill height above toe; prefer absolute elevation, fall back to legacy height."""
    if not (req.backfill_include_pressure or req.backfill_include_weight):
        return 0.0
    if getattr(req, "backfill_elevation", 0) and req.backfill_elevation > 0:
        h = req.backfill_elevation - geom.toe_elevation
    else:
        h = req.backfill_height
    return max(h, 0.0)


@app.post("/calculate", response_model=CalculationResponse)
def calculate(req: LoadCaseRequest):
    try:
        # ── Build geometry ────────────────────────────────────────────
        coords = [(p.x, p.y) for p in req.coordinates]
        utp    = (req.upstream_top_point.x, req.upstream_top_point.y)
        geom   = DamGeometry(coordinates=coords, upstream_top_point=utp)

        # Resolve silt / backfill heights from absolute elevation (preferred)
        silt_h = _resolve_silt_height(req, geom)
        bf_h   = _resolve_backfill_height(req, geom)

        # ── Build configs ─────────────────────────────────────────────
        mat = MaterialProperties(
            unit_weight_dam   = req.unit_weight_dam,
            unit_weight_water = req.unit_weight_water,
            friction_coeff    = req.friction_coeff,
        )
        drainage = DrainageConfig(
            include            = req.drainage_include,
            distance_from_heel = req.drainage_distance_from_heel,
            reduction_factor   = req.drainage_reduction_factor,
        )
        silt = SiltConfig(
            include               = req.silt_include,
            height_us             = silt_h,
            unit_weight_submerged = req.silt_unit_weight_submerged,
            phi_deg               = req.silt_phi_deg,
        )
        backfill = BackfillConfig(
            include_pressure      = req.backfill_include_pressure,
            include_weight        = req.backfill_include_weight,
            height                = bf_h,
            coeff_pressure        = req.backfill_coeff_pressure,
            unit_weight_dry       = req.backfill_unit_weight_dry,
            unit_weight_wet       = req.backfill_unit_weight_wet,
            unit_weight_submerged = req.backfill_unit_weight_submerged,
        )
        ice_base    = IcePressureConfig(include=req.ice_include, pressure=req.ice_pressure)
        rock_bolt   = RockBoltConfig(include=req.rock_bolt_include,
                                     force_per_m=req.rock_bolt_force_per_m,
                                     cover_from_heel=req.rock_bolt_cover_from_heel)
        rock_anchor = RockAnchorConfig(include=req.rock_anchor_include,
                                       force_per_m=req.rock_anchor_force_per_m,
                                       cover_from_heel=req.rock_anchor_cover_from_heel)
        earthquake  = EarthquakeConfig(
            include=req.run_EQ,
            a_h=req.eq_a_h,
            a_v=req.eq_a_v)
        applied     = AppliedForceConfig(
            vertical_forces   = [tuple(v) for v in req.applied_vertical_forces],
            horizontal_forces = [tuple(h) for h in req.applied_horizontal_forces],
        )

        # ── Warnings ──────────────────────────────────────────────────
        warnings = []
        hrv_head = req.HRV_us - geom.heel_elevation   # water depth above heel at HRV
        if (rock_bolt.include and req.rock_bolt_apply_depth_limit
                and hrv_head > req.rock_bolt_depth_limit):
            warnings.append(
                f"Rock bolts disabled: HRV water depth above heel "
                f"{hrv_head:.2f} m exceeds the {req.rock_bolt_depth_limit:.1f} m "
                f"depth limit for rock bolt use.")

        # ── Which cases to run ────────────────────────────────────────
        cases = []
        if req.run_HRV:
            cases.append(('HRV',               req.HRV_us, req.HRV_ds, True))
        if req.run_DFV:
            cases.append(('DFV',               req.DFV_us, req.DFV_ds, True))
        if req.run_MFV:
            cases.append(('MFV',               req.MFV_us, req.MFV_ds, True))
        if req.run_DFV_no_bolts:
            cases.append(('DFV (no rock bolts)', req.DFV_us, req.DFV_ds, False))
        if req.run_EQ and (req.eq_a_h > 0 or req.eq_a_v > 0):
            cases.append(('HRV+EQ (X-dom)', req.HRV_us, req.HRV_ds, True))
            cases.append(('HRV+EQ (Y-dom)', req.HRV_us, req.HRV_ds, True))

        if not cases:
            raise HTTPException(status_code=400, detail="No load cases selected.")

        # ── Run each case ─────────────────────────────────────────────
        results_out = []
        for case_name, wl_us, wl_ds, incl_rb in cases:
            # Ice only for HRV
            ice_case = IcePressureConfig(
                include  = req.ice_include and (case_name == 'HRV'),
                pressure = req.ice_pressure,
            )
            # EQ cases: scale accelerations per Eurocode 8 combination.
            # For all non-EQ cases, earthquake is always disabled regardless
            # of the run_EQ checkbox — earthquake forces must NEVER appear
            # in HRV, DFV, MFV, or DFV (no rock bolts).
            if case_name == 'HRV+EQ (X-dom)' and earthquake.include:
                eq_for_case = EarthquakeConfig(
                    include=True, a_h=earthquake.a_h, a_v=0.3*earthquake.a_v)
            elif case_name == 'HRV+EQ (Y-dom)' and earthquake.include:
                eq_for_case = EarthquakeConfig(
                    include=True, a_h=0.3*earthquake.a_h, a_v=earthquake.a_v)
            else:
                eq_for_case = EarthquakeConfig(include=False)

            # EQ cases exclude ice per Eurocode 8
            ice_for_case = IcePressureConfig(include=False) \
                if 'EQ' in case_name else ice_case

            res = run_load_case(
                case_name=case_name, geom=geom, mat=mat,
                wl_us=wl_us, wl_ds=wl_ds,
                drainage=drainage, silt=silt, backfill=backfill,
                ice=ice_for_case, rock_bolt=rock_bolt, rock_anchor=rock_anchor,
                applied=applied, include_rock_bolts=incl_rb,
                rb_depth_limit_apply=req.rock_bolt_apply_depth_limit,
                rb_depth_limit=req.rock_bolt_depth_limit,
                earthquake=eq_for_case,
                fs_uls=req.fs_uls, fs_als=req.fs_als,
                res_uls=req.res_uls, res_als=req.res_als,
            )
            res['earthquake'] = eq_for_case   # store for plotting
            # Generate plots (returns {'dam':..., 'stress':...})
            plot_imgs = plot_to_base64(res, geom, mat, drainage, silt, backfill)

            # Serialize forces
            forces_out = [
                ForceRow(
                    name        = r['name'],
                    V           = round(r['V'],       3),
                    H           = round(r['H'],       3),
                    x_from_toe  = round(r['x_from_toe'], 4),
                    y_from_toe  = round(r['y_from_toe'], 4),
                    M_res       = round(r['M_res'],   3),
                    M_ov        = round(r['M_ov'],    3),
                    stabilising = r['stabilising'],
                )
                for r in res['rows']
            ]

            def safe(v):
                """Replace inf with a large sentinel for JSON serialisation."""
                if v == float('inf'):  return 9999.0
                if v == float('-inf'): return -9999.0
                return round(v, 4)

            results_out.append(CaseResult(
                case_name            = res['case_name'],
                sum_V                = safe(res['sum_V']),
                H_net                = safe(res['H_net']),
                sum_M_res            = safe(res['sum_M_res']),
                sum_M_ov             = safe(res['sum_M_ov']),
                x_resultant          = safe(res['x_resultant']),
                eccentricity         = safe(res['eccentricity']),
                in_middle_third      = res['in_middle_third'],
                resultant_check_type = res['resultant_check_type'],
                fs_threshold         = res.get('fs_threshold', 1.5),
                sigma_toe            = safe(res['sigma_toe']),
                sigma_heel           = safe(res['sigma_heel']),
                FS_sliding           = safe(res['FS_sliding']),
                FS_overturning       = safe(res['FS_overturning']),
                tension_length       = safe(res['tension_length']),
                forces               = forces_out,
                messages             = res.get('messages', []),
                plot_dam_base64      = plot_imgs['dam'],
                plot_stress_base64   = plot_imgs['stress'],
            ))

        geom_info = GeometryInfo(
            toe_elevation          = geom.toe_elevation,
            heel_elevation         = geom.heel_elevation,
            base_length_horizontal = round(geom.base_length_horizontal, 3),
            dam_height             = round(geom.dam_top_elevation_rel,  3),
            warnings               = warnings,
        )
        return CalculationResponse(geometry=geom_info, results=results_out)

    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception:
        raise HTTPException(status_code=500, detail=traceback.format_exc())


# =============================================================================
# EXCEL EXPORT ENDPOINT
# =============================================================================

class ExportRequest(BaseModel):
    """Same calculation response data sent back from the frontend for export."""
    geometry: GeometryInfo
    results:  List[CaseResult]
    params:   dict = {}   # input parameters for the cover block


@app.post("/export")
def export_excel(data: ExportRequest):
    """
    Build an Excel workbook: a Summary sheet plus ONE sheet for the dam height,
    with all load cases stacked vertically. Two figures (dam + stress) are placed
    in a far-right column so they never overlap the tables/text.
    """
    try:
        wb = _build_workbook(
            title="GRAVITY DAM STABILITY — RESULTS",
            blocks=[_HeightBlock(
                label="Main Dam",
                heel_elevation=data.geometry.heel_elevation,
                toe_elevation=data.geometry.toe_elevation,
                base_width=data.geometry.base_length_horizontal,
                dam_height=data.geometry.dam_height,
                results=data.results,
            )],
            params=data.params,
        )
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(
            buf,
            media_type=("application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet"),
            headers={"Content-Disposition":
                     'attachment; filename="dam_stability_results.xlsx"'},
        )
    except Exception:
        raise HTTPException(status_code=500, detail=traceback.format_exc())


# =============================================================================
# SHARED EXCEL BUILDER  (one sheet per dam height, load cases stacked)
# =============================================================================

from dataclasses import dataclass as _dataclass

@_dataclass
class _HeightBlock:
    """One dam height = one worksheet, holding all its load-case results."""
    label:          str
    heel_elevation: float
    toe_elevation:  float
    base_width:     float
    dam_height:     float
    results:        list   # list of CaseResult-like objects


def _build_workbook(title: str, blocks: list, params: dict = {}):
    """
    Construct the workbook for printing as an A4-portrait report attachment.

    Layout per dam height (one worksheet), vertical order for each load case:
        1. Load-case banner
        2. Warnings / error messages
        3. Force table
        4. Factor-of-safety / resultant / stress summary table
           (with the diagram placed beside it if it fits on A4, else below)

    Fonts: Aptos 9 pt body; larger sizes for headings.
    Column grid is sized so the printable content fits A4 portrait width.
    """
    BODY_FONT  = "Aptos"
    BODY_SZ    = 9
    H1_SZ      = 14       # workbook / sheet title
    H2_SZ      = 11       # load-case banner
    H3_SZ      = 10       # sub-headers

    HEADER_FILL = "EEEEEE"   # light gray for all headers — only colour used
    WHITE       = "FFFFFF"
    BLUE_DARK = BLUE_MID = BLUE_LIGHT = HEADER_FILL
    GREEN_DARK = RED_DARK = "000000"
    GREEN_LIGHT = RED_LIGHT = GRAY_LIGHT = None

    def dcn(n):
        """Display case name: internal 'HRV' → 'HRV+IS' in Excel output."""
        return 'HRV+IS' if n == 'HRV' else n

    def hfill(c): return PatternFill("solid", fgColor=c) if c else None
    def _side():  return Side(style="thin", color="AAAAAA")
    tbord = Border(left=_side(), right=_side(), top=_side(), bottom=_side())

    def wc(ws, r, c, val, bold=False, fill=None, align='left',
           color="000000", sz=BODY_SZ, wrap=False, border=True):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(bold=bold, name=BODY_FONT, size=sz, color=color)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if border: cell.border = tbord
        if fill is not None: cell.fill = fill
        return cell

    def nf(ws, r, c, val, fmt='0.00', bold=False, fill=None, align='right'):
        cell = ws.cell(row=r, column=c, value=val)
        cell.number_format = fmt
        cell.font = Font(bold=bold, name=BODY_FONT, size=BODY_SZ)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = tbord
        if fill is not None: cell.fill = fill
        return cell

    wb = Workbook(); wb.remove(wb.active)

    # ── A4 portrait grid ──────────────────────────────────────────────
    # 12 content columns A–L. A4 portrait printable width ≈ 18 cm.
    # Excel width unit ≈ 7 px ≈ 0.19 cm. Keep total ≈ 95 units ≈ 18 cm.
    # Force table : columns A–F.
    # FS table    : columns A–C (placed below the force table).
    # Diagram     : columns H–L (beside the FS table) if it fits, else below.
    N_COLS       = 8                      # A..H
    LAST_COL     = get_column_letter(N_COLS)
    IMG_COL      = "A"                    # diagram anchor column (always below)

    def setup_page(ws):
        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize   = 9          # A4
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5,
                                      header=0.2, footer=0.2)

    # =====================================================================
    # SUMMARY SHEET
    # =====================================================================
    ws = wb.create_sheet("Summary")
    setup_page(ws)
    ws.merge_cells(f"A1:{LAST_COL}1")
    t = ws["A1"]; t.value = title
    t.font = Font(bold=True, size=H1_SZ, color="000000", name=BODY_FONT)
    t.fill = hfill(BLUE_DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # ── Input parameters block ────────────────────────────────────────
    row_p = 2   # start row for params
    if params:
        def psec(ws, r, heading):
            ws.merge_cells(f"A{r}:{LAST_COL}{r}")
            c = ws[f"A{r}"]
            c.value = heading
            c.font  = Font(bold=True, size=BODY_SZ, name=BODY_FONT)
            c.fill  = hfill(HEADER_FILL)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[r].height = 14
            return r + 1

        def prow(ws, r, label, value, unit=''):
            ws.cell(row=r, column=1, value=label).font = Font(name=BODY_FONT, size=BODY_SZ)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", indent=2)
            val_str = f"{value}  {unit}".strip() if unit else str(value)
            ws.cell(row=r, column=2, value=val_str).font = Font(name=BODY_FONT, size=BODY_SZ, bold=True)
            ws.row_dimensions[r].height = 13
            return r + 1

        g  = params.get
        r  = row_p
        r  = psec(ws, r, "Geometry")
        r  = prow(ws, r, "Toe elevation",   g('toe_elevation','—'),    'm')
        r  = prow(ws, r, "Heel elevation",  g('heel_elevation','—'),   'm')
        r  = prow(ws, r, "Base length",     g('base_length','—'),      'm')
        r  = prow(ws, r, "Dam height",      g('dam_height','—'),       'm')
        r  = psec(ws, r, "Material Properties")
        r  = prow(ws, r, "Unit weight (dam)",  g('unit_weight_dam', '—'), 'kN/m³')
        r  = prow(ws, r, "Unit weight (water)",g('unit_weight_water','—'),'kN/m³')
        r  = prow(ws, r, "Friction coefficient", g('friction_coeff','—'), '')
        r  = psec(ws, r, "Water Levels")
        r  = prow(ws, r, "HRV+IS upstream WL",  g('HRV_us','—'), 'm')
        r  = prow(ws, r, "HRV+IS downstream WL", g('HRV_ds','—'), 'm')
        r  = prow(ws, r, "DFV upstream WL",   g('DFV_us','—'), 'm')
        r  = prow(ws, r, "DFV downstream WL", g('DFV_ds','—'), 'm')
        r  = prow(ws, r, "MFV upstream WL",   g('MFV_us','—'), 'm')
        r  = prow(ws, r, "MFV downstream WL", g('MFV_ds','—'), 'm')
        r  = psec(ws, r, "Acceptance Criteria")
        r  = prow(ws, r, "FS (ULS — HRV, DFV)", g('fs_uls', 1.5), '')
        r  = prow(ws, r, "FS (ALS — MFV, EQ)",  g('fs_als', 1.1), '')
        r  = prow(ws, r, "Resultant zone (ULS)", g('res_uls','middle_third').replace('_',' '), '')
        r  = prow(ws, r, "Resultant zone (ALS)", g('res_als','l6').replace('_',' ').replace('l6','L/6'), '')
        if g('drainage_include', False):
            r = psec(ws, r, "Drainage Curtain")
            r = prow(ws, r, "Distance from heel", g('drainage_dist','—'), 'm')
            r = prow(ws, r, "Reduction factor",   g('drainage_rf','—'),   '')
        if g('silt_include', False):
            r = psec(ws, r, "Silt")
            r = prow(ws, r, "Silt surface elevation", g('silt_elevation','—'), 'm')
            r = prow(ws, r, "Submerged unit weight",  g('silt_unit_weight','—'), 'kN/m³')
            r = prow(ws, r, "Friction angle φ",       g('silt_phi','—'), '°')
        if g('ice_include', False):
            r = psec(ws, r, "Ice Pressure")
            r = prow(ws, r, "Ice pressure", g('ice_pressure','—'), 'kN/m')
        if g('rock_bolt_include', False):
            r = psec(ws, r, "Rock Bolts")
            r = prow(ws, r, "Force per unit width", g('rock_bolt_force','—'), 'kN/m')
            r = prow(ws, r, "Cover from heel",      g('rock_bolt_cover','—'), 'm')
        if g('rock_anchor_include', False):
            r = psec(ws, r, "Rock Anchors")
            r = prow(ws, r, "Force per unit width", g('rock_anchor_force','—'), 'kN/m')
            r = prow(ws, r, "Cover from heel",      g('rock_anchor_cover','—'), 'm')
        if g('run_EQ', False):
            r = psec(ws, r, "Earthquake")
            r = prow(ws, r, "Horizontal acceleration aₕ/g", g('eq_ah','—'), '')
            r = prow(ws, r, "Vertical acceleration aᵥ/g",   g('eq_av','—'), '')
        # Blank separator row
        ws.row_dimensions[r].height = 8; r += 1
        row_p = r
    else:
        row_p = 2

    hdrs = ["Dam Height", "Load Case", "FS Sliding", "FS Overturning",
            "Resultant", "x_res (m)", "e (m)", "σ_toe", "σ_heel", "Tens.(m)"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=row_p, column=ci, value=h)
        c.font = Font(bold=True, color="000000", name=BODY_FONT, size=BODY_SZ)
        c.fill = hfill(BLUE_MID)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = tbord
    ws.row_dimensions[row_p].height = 30

    r_idx = row_p + 1
    for blk in blocks:
        for res in blk.results:
            fs_ok  = res.FS_sliding >= res.fs_threshold
            mid_ok = res.in_middle_third
            wc(ws, r_idx, 1, blk.label, bold=True)
            wc(ws, r_idx, 2, dcn(res.case_name), align='center')

            def fs_cell(col, val, ok):
                disp = "∞" if val >= 9998 else val
                cc = ws.cell(row=r_idx, column=col, value=disp)
                if not isinstance(disp, str): cc.number_format = '0.000'
                cc.font = Font(bold=True, name=BODY_FONT, size=BODY_SZ)
                cc.alignment = Alignment(horizontal="center", vertical="center")
                cc.border = tbord
                if ok is True:  cc.fill = hfill("C6EFCE")
                elif ok is False: cc.fill = hfill("FFC7CE")
            fs_cell(3, res.FS_sliding, fs_ok)
            fs_cell(4, res.FS_overturning, True)

            ck = ws.cell(row=r_idx, column=5, value="✓ OK" if mid_ok else "✗ OUT")
            ck.font = Font(bold=True, name=BODY_FONT, size=BODY_SZ)
            ck.alignment = Alignment(horizontal="center", vertical="center")
            ck.border = tbord
            ck.fill = hfill("C6EFCE") if mid_ok else hfill("FFC7CE")
            xr_cell = ws.cell(row=r_idx, column=6, value=res.x_resultant)
            xr_cell.number_format = '0.000'
            xr_cell.font = Font(name=BODY_FONT, size=BODY_SZ)
            xr_cell.alignment = Alignment(horizontal="center", vertical="center")
            xr_cell.border = tbord
            xr_cell.fill = hfill("C6EFCE") if mid_ok else hfill("FFC7CE")
            nf(ws, r_idx, 7, res.eccentricity, '0.000')
            nf(ws, r_idx, 8, res.sigma_toe, '0.0')
            nf(ws, r_idx, 9, res.sigma_heel, '0.0')
            nf(ws, r_idx, 10, res.tension_length if res.tension_length > 0.001 else 0,
               '0.000')
            ws.row_dimensions[r_idx].height = 15
            r_idx += 1

    for ci, w in enumerate([13,10,9,11,9,8,7,7,8,8], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # =====================================================================
    # ONE SHEET PER DAM HEIGHT
    # =====================================================================
    for blk in blocks:
        ws = wb.create_sheet(blk.label[:31])
        setup_page(ws)

        # Column widths tuned to A4 portrait (sum of A..L ≈ 92 units ≈ 18 cm)
        # Column layout: A–H = force table (8 cols), A–D reused by FS table below.
        # Col D must be wide enough for the "Note / Range" text in the FS table.
        # Total A–H ≈ 85 units, fits A4 portrait width.
        col_w = {'A':17,'B':8,'C':8,'D':22,'E':8,'F':8,'G':9,'H':9}
        for col, w in col_w.items():
            ws.column_dimensions[col].width = w

        # Sheet title
        ws.merge_cells(f"A1:{LAST_COL}1")
        t2 = ws["A1"]
        t2.value = f"Dam Height: {blk.label}"
        t2.font = Font(bold=True, size=H1_SZ, color="000000", name=BODY_FONT)
        t2.fill = hfill(BLUE_DARK)
        t2.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        ws.merge_cells(f"A2:{LAST_COL}2")
        g = ws["A2"]
        g.value = (f"Heel: {blk.heel_elevation:.2f} m    |    "
                   f"Toe: {blk.toe_elevation:.2f} m    |    "
                   f"Base width L: {blk.base_width:.2f} m    |    "
                   f"Dam height: {blk.dam_height:.2f} m")
        g.font = Font(size=BODY_SZ, italic=True, name=BODY_FONT)
        g.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 16

        row = 4

        for res in blk.results:
            # ── 1. Load-case banner ───────────────────────────────────
            ws.merge_cells(f"A{row}:{LAST_COL}{row}")
            cb = ws[f"A{row}"]
            cb.value = f"LOAD CASE: {dcn(res.case_name)}"
            cb.font = Font(bold=True, size=H2_SZ, color="000000", name=BODY_FONT)
            cb.fill = hfill(BLUE_MID)
            cb.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row].height = 20
            row += 1

            # ── 2. Warnings / messages ────────────────────────────────
            if res.messages:
                TYPE_ICON  = {"info":"ℹ","warning":"⚠","alert":"⛔"}
                TYPE_FILL  = {"info":"D6E8F7","warning":"FFF3CD","alert":"FAD7D7"}
                TYPE_COLOR = {"info":"1A3A5C","warning":"7A5000","alert":"8B1A1A"}
                for m in res.messages:
                    mtype = m.get("type","info") if isinstance(m, dict) else "info"
                    mtext = m.get("text",str(m)) if isinstance(m, dict) else str(m)
                    ws.merge_cells(f"A{row}:{LAST_COL}{row}")
                    mc = ws[f"A{row}"]
                    mc.value = f"{TYPE_ICON.get(mtype,'ℹ')}  {mtext}"
                    mc.font = Font(name=BODY_FONT, size=BODY_SZ, color=TYPE_COLOR.get(mtype,"1A3A5C"))
                    mc.fill = hfill(TYPE_FILL.get(mtype,"D6E8F7"))
                    mc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    mc.border = tbord
                    ws.row_dimensions[row].height = 26
                    row += 1
                row += 1   # spacer

            # ── 3. Force table (columns A–F) ──────────────────────────
            fhdr = ["Force","Stab/Dest","V (kN/m)","H (kN/m)","x_arm (m)","y_arm (m)","M_res","M_ov"]
            for ci, h in enumerate(fhdr, 1):
                c = ws.cell(row=row, column=ci, value=h)
                c.font = Font(bold=True, color="000000", name=BODY_FONT, size=BODY_SZ)
                c.fill = hfill(BLUE_MID)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = tbord
            ws.row_dimensions[row].height = 24
            row += 1
            for f in res.forces:
                wc(ws, row, 1, f.name)
                wc(ws, row, 2, "Stab" if f.stabilising else "Dest",
                   align='center', bold=True)
                nf(ws, row, 3, f.V,          '0.00')
                nf(ws, row, 4, f.H,          '0.00')
                nf(ws, row, 5, f.x_from_toe, '0.000')
                nf(ws, row, 6, f.y_from_toe, '0.000')
                nf(ws, row, 7, f.M_res,      '0.00')
                nf(ws, row, 8, f.M_ov,       '0.00')
                ws.row_dimensions[row].height = 13
                row += 1
            # Resultant totals row (kept in force table only)
            wc(ws, row, 1, "RESULTANT", bold=True)
            wc(ws, row, 2, "")
            nf(ws, row, 3, res.sum_V,   '0.00', bold=True)
            nf(ws, row, 4, res.H_net,   '0.00', bold=True)
            wc(ws, row, 5, "")
            wc(ws, row, 6, "")
            nf(ws, row, 7, res.sum_M_res,'0.00', bold=True)
            nf(ws, row, 8, res.sum_M_ov, '0.00', bold=True)
            ws.row_dimensions[row].height = 14
            row += 2   # spacer

            # ── 4. FS / resultant summary table (cols A–C) ────────────
            #    NO sum of forces/moments here (those live in the force table).
            fs_table_top = row
            hdr2 = ["Result", "Value", "Status", "Note / Range"]
            for ci, h in enumerate(hdr2, 1):
                c = ws.cell(row=row, column=ci, value=h)
                c.font = Font(bold=True, color="000000", name=BODY_FONT, size=BODY_SZ)
                c.fill = hfill(BLUE_MID)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = tbord
            ws.row_dimensions[row].height = 16
            row += 1

            fs_thr = res.fs_threshold
            L = res.x_resultant / (1 - res.eccentricity / (res.x_resultant or 1))                 if False else None  # compute L from geometry info stored in block
            # Compute L (base width) from the block context
            blk_L = blk.base_width   # metres

            if res.resultant_check_type == "L/6–5L/6":
                x_lo = round(blk_L / 6, 3); x_hi = round(5 * blk_L / 6, 3)
                x_range_str = f"L/6 – 5L/6  ({x_lo:.3f} – {x_hi:.3f} m)"
                crit_str    = f"L/6 – 5L/6 (L={blk_L:.2f} m)"
            else:
                x_lo = round(blk_L / 3, 3); x_hi = round(2 * blk_L / 3, 3)
                x_range_str = f"L/3 – 2L/3  ({x_lo:.3f} – {x_hi:.3f} m)"
                crit_str    = f"L/3 – 2L/3 (L={blk_L:.2f} m)"

            def status_txt(ok): return "OK" if ok else "CHECK"
            # items: (label, value, ok_flag, number_format, note_text)
            items = [
                ("FS Sliding",      res.FS_sliding,      res.FS_sliding >= fs_thr, '0.000', f"Threshold ≥ {fs_thr:.1f}"),
                ("FS Overturning",  res.FS_overturning,  None, '0.000', ""),
                ("Resultant",       "✓ OK" if res.in_middle_third else "✗ OUTSIDE",
                                     res.in_middle_third, '@', crit_str),
                ("x_resultant (m)", res.x_resultant,     res.in_middle_third, '0.000', x_range_str),
                ("Eccentricity (m)",res.eccentricity,    None, '0.000', ""),
                ("σ_toe (kN/m²)",   res.sigma_toe,       None, '0.00', "Compression = positive"),
                ("σ_heel (kN/m²)",  res.sigma_heel,      None, '0.00', "Compression = positive"),
                ("Tension len (m)", res.tension_length,  None, '0.000', ""),
            ]
            for (lbl, val, ok, fmt, note) in items:
                wc(ws, row, 1, lbl, bold=True)
                cc = ws.cell(row=row, column=2, value=val)
                if fmt != '@':
                    cc.number_format = fmt
                    if isinstance(val, (int,float)) and val >= 9998: cc.value = "∞"
                cc.font = Font(bold=True, name=BODY_FONT, size=BODY_SZ)
                cc.alignment = Alignment(horizontal="center", vertical="center")
                cc.border = tbord
                if ok is True:  cc.fill = hfill("C6EFCE")
                elif ok is False: cc.fill = hfill("FFC7CE")
                # Status column
                if ok is None:
                    wc(ws, row, 3, "—", align='center')
                else:
                    st_cell = ws.cell(row=row, column=3, value=status_txt(ok))
                    st_cell.font = Font(bold=True, name=BODY_FONT, size=BODY_SZ)
                    st_cell.alignment = Alignment(horizontal="center", vertical="center")
                    st_cell.border = tbord
                    st_cell.fill = hfill("C6EFCE") if ok else hfill("FFC7CE")
                # Note column
                wc(ws, row, 4, note, wrap=True, sz=BODY_SZ-1)
                ws.row_dimensions[row].height = 14
                row += 1
            fs_table_bottom = row - 1

            # ── Diagram placement ─────────────────────────────────────
            # Place dam figure below the FS table.
            dam_b64    = getattr(res, "plot_dam_base64", "") or ""
            target_w   = 500      # px — larger figure for readability

            dam_rows = 0
            im = None
            if dam_b64:
                im = XLImage(io.BytesIO(base64.b64decode(dam_b64)))
                im.height = int(target_w * im.height / im.width) if im.width else 300
                im.width  = target_w
                dam_rows = int(im.height / 18) + 1

            below = fs_table_bottom + 2
            if im: ws.add_image(im, f"A{below}")
            row = below + dam_rows + 2

    return wb


# =============================================================================
# MULTI-HEIGHT ANALYSIS
# =============================================================================

class SectionInput(BaseModel):
    label:     str
    us_toe_y:  float
    ds_toe_y:  float
    hrv_ds_wl: float
    dfv_ds_wl: float
    mfv_ds_wl: float

class MultiHeightRequest(BaseModel):
    base_request: LoadCaseRequest
    sections:     List[SectionInput]

class SectionResponse(BaseModel):
    label:          str
    heel_elevation: float
    toe_elevation:  float
    base_width:     float
    dam_height:     float
    results:        List[CaseResult]

class MultiHeightResponse(BaseModel):
    sections: List[SectionResponse]


def _safe(v):
    if v == float('inf'):  return 9999.0
    if v == float('-inf'): return -9999.0
    return round(v, 4)


def _solve_case(case_name, geom, mat, wl_us, wl_ds,
                drainage, silt, backfill, ice_pressure, ice_on,
                rock_bolt, rock_anchor, applied, incl_rb,
                rb_depth_limit_apply=True, rb_depth_limit=7.0,
                earthquake=None):
    """Run one load case and return a fully-populated CaseResult."""
    ice_case = IcePressureConfig(include=(ice_on and case_name == 'HRV'),
                                 pressure=ice_pressure)
    res = run_load_case(
        case_name=case_name, geom=geom, mat=mat,
        wl_us=wl_us, wl_ds=wl_ds,
        drainage=drainage, silt=silt, backfill=backfill,
        ice=ice_case, rock_bolt=rock_bolt, rock_anchor=rock_anchor,
        applied=applied, include_rock_bolts=incl_rb,
        rb_depth_limit_apply=rb_depth_limit_apply,
        rb_depth_limit=rb_depth_limit,
        earthquake=earthquake if earthquake is not None else EarthquakeConfig(include=False),
    )
    res['earthquake'] = earthquake   # store for plotting
    imgs = plot_to_base64(res, geom, mat, drainage, silt, backfill)
    forces_out = [
        ForceRow(name=r['name'], V=round(r['V'],3), H=round(r['H'],3),
                 x_from_toe=round(r['x_from_toe'],4), y_from_toe=round(r['y_from_toe'],4),
                 M_res=round(r['M_res'],3), M_ov=round(r['M_ov'],3),
                 stabilising=r['stabilising'])
        for r in res['rows']
    ]
    return CaseResult(
        case_name=res['case_name'],
        sum_V=_safe(res['sum_V']), H_net=_safe(res['H_net']),
        sum_M_res=_safe(res['sum_M_res']), sum_M_ov=_safe(res['sum_M_ov']),
        x_resultant=_safe(res['x_resultant']), eccentricity=_safe(res['eccentricity']),
        in_middle_third=res['in_middle_third'],
        resultant_check_type=res['resultant_check_type'],
        fs_threshold=res.get('fs_threshold', 1.5),
        sigma_toe=_safe(res['sigma_toe']), sigma_heel=_safe(res['sigma_heel']),
        FS_sliding=_safe(res['FS_sliding']), FS_overturning=_safe(res['FS_overturning']),
        tension_length=_safe(res['tension_length']),
        forces=forces_out, messages=res.get('messages', []),
        plot_dam_base64=imgs['dam'], plot_stress_base64=imgs['stress'],
    )


def _run_section(req: LoadCaseRequest, sec: SectionInput) -> SectionResponse:
    """
    Re-run stability at a lower cross-section, keeping the SAME dam face
    geometry above the new base. New heel/toe x are interpolated on the
    original faces at the requested elevations; a straight base joins them.
    """
    orig_coords = [(p.x, p.y) for p in req.coordinates]
    utp_orig    = (req.upstream_top_point.x, req.upstream_top_point.y)

    # Heel/toe: same rule as DamGeometry.
    # x ≤ utp_x → upstream face (heel = lowest y there)
    # x >  utp_x → downstream face (toe = lowest y there)
    n = len(orig_coords)

    def nidx(pt):
        return min(range(len(orig_coords)),
                   key=lambda i: (orig_coords[i][0]-pt[0])**2+(orig_coords[i][1]-pt[1])**2)

    us_pts = [(x, y) for x, y in orig_coords if x <= utp_orig[0]]
    ds_pts = [(x, y) for x, y in orig_coords if x >  utp_orig[0]]
    heel_pt = min(us_pts, key=lambda p: p[1])
    toe_pt  = min(ds_pts, key=lambda p: p[1])

    heel_idx, toe_idx = nidx(heel_pt), nidx(toe_pt)
    utp_idx = nidx(utp_orig)
    top_y = max(y for _, y in orig_coords)
    crest_pts = [(x, y) for x, y in orig_coords if abs(y-top_y) < 1e-6]
    ds_crest = min(crest_pts, key=lambda p: abs(p[0]-toe_pt[0]))
    ds_top_idx = nidx(ds_crest)

    def walk(s, e, d):
        path=[]; i=s
        for _ in range(n+1):
            path.append(orig_coords[i])
            if i==e: break
            i=(i+d)%n
        return path
    def best_face(s, e):
        cands=[walk(s,e,d) for d in (1,-1)]
        cands=[p for p in cands if p[-1][1] >= p[0][1]]
        return min(cands, key=len) if cands else walk(s,e,1)

    us_face = best_face(heel_idx, utp_idx)
    ds_face = best_face(toe_idx,  ds_top_idx)

    y_heel_new, y_toe_new = sec.us_toe_y, sec.ds_toe_y
    if y_heel_new < heel_pt[1]-1e-3:
        raise ValueError(f"Section '{sec.label}': new heel elev ({y_heel_new}) below original heel ({heel_pt[1]}).")
    if y_toe_new < toe_pt[1]-1e-3:
        raise ValueError(f"Section '{sec.label}': new toe elev ({y_toe_new}) below original toe ({toe_pt[1]}).")
    if y_heel_new >= top_y-1e-3 or y_toe_new >= top_y-1e-3:
        raise ValueError(f"Section '{sec.label}': new base at/above crest ({top_y}).")

    def x_on_face(face, yt):
        for i in range(len(face)-1):
            x0,y0=face[i]; x1,y1=face[i+1]
            lo,hi=min(y0,y1),max(y0,y1)
            if lo-1e-6 <= yt <= hi+1e-6:
                if abs(y1-y0)<1e-9: return (x0+x1)/2
                return x0+(yt-y0)/(y1-y0)*(x1-x0)
        return face[0][0]

    new_heel = (x_on_face(us_face, y_heel_new), y_heel_new)
    new_toe  = (x_on_face(ds_face, y_toe_new),  y_toe_new)

    def aeq(a,b): return abs(a[0]-b[0])<1e-6 and abs(a[1]-b[1])<1e-6
    def push(lst,pt):
        if not lst or not aeq(lst[-1],pt): lst.append(pt)

    new_coords=[new_heel]
    for pt in us_face:
        if pt[1] > y_heel_new+1e-6: push(new_coords, pt)
    i=utp_idx
    for _ in range(n+1):
        push(new_coords, orig_coords[i])
        if i==ds_top_idx: break
        i=(i+1)%n
    for pt in reversed(ds_face):
        if pt[1] > y_toe_new+1e-6: push(new_coords, pt)
    push(new_coords, new_toe)

    if len(new_coords) < 3:
        raise ValueError(f"Section '{sec.label}': polygon < 3 vertices after clipping.")

    geom = DamGeometry(coordinates=new_coords, upstream_top_point=utp_orig)

    mat = MaterialProperties(
        unit_weight_dam=req.unit_weight_dam,
        unit_weight_water=req.unit_weight_water,
        friction_coeff=req.friction_coeff)
    drainage = DrainageConfig(
        include=req.drainage_include,
        distance_from_heel=req.drainage_distance_from_heel,
        reduction_factor=req.drainage_reduction_factor)
    silt = SiltConfig(
        include=req.silt_include,
        height_us=_resolve_silt_height(req, geom),
        unit_weight_submerged=req.silt_unit_weight_submerged,
        phi_deg=req.silt_phi_deg)
    backfill = BackfillConfig(
        include_pressure=req.backfill_include_pressure,
        include_weight=req.backfill_include_weight,
        height=_resolve_backfill_height(req, geom),
        coeff_pressure=req.backfill_coeff_pressure,
        unit_weight_dry=req.backfill_unit_weight_dry,
        unit_weight_wet=req.backfill_unit_weight_wet,
        unit_weight_submerged=req.backfill_unit_weight_submerged)
    rock_bolt = RockBoltConfig(
        include=req.rock_bolt_include,
        force_per_m=req.rock_bolt_force_per_m,
        cover_from_heel=req.rock_bolt_cover_from_heel)
    rock_anchor = RockAnchorConfig(
        include=req.rock_anchor_include,
        force_per_m=req.rock_anchor_force_per_m,
        cover_from_heel=req.rock_anchor_cover_from_heel)
    applied = AppliedForceConfig(
        vertical_forces=[tuple(v) for v in req.applied_vertical_forces],
        horizontal_forces=[tuple(h) for h in req.applied_horizontal_forces])

    cases = []
    if req.run_HRV: cases.append(('HRV', req.HRV_us, sec.hrv_ds_wl, True))
    if req.run_DFV: cases.append(('DFV', req.DFV_us, sec.dfv_ds_wl, True))
    if req.run_MFV: cases.append(('MFV', req.MFV_us, sec.mfv_ds_wl, True))
    if req.run_DFV_no_bolts:
        cases.append(('DFV (no rock bolts)', req.DFV_us, sec.dfv_ds_wl, False))
    if req.run_EQ and (req.eq_a_h > 0 or req.eq_a_v > 0):
        cases.append(('HRV+EQ (X-dom)', req.HRV_us, sec.hrv_ds_wl, True))
        cases.append(('HRV+EQ (Y-dom)', req.HRV_us, sec.hrv_ds_wl, True))

    # Build earthquake configs — same Eurocode 8 combination rule as main calculate()
    base_eq = EarthquakeConfig(include=req.run_EQ, a_h=req.eq_a_h, a_v=req.eq_a_v)

    def eq_for(case_name):
        if case_name == 'HRV+EQ (X-dom)' and base_eq.include:
            return EarthquakeConfig(include=True, a_h=base_eq.a_h, a_v=0.3*base_eq.a_v)
        elif case_name == 'HRV+EQ (Y-dom)' and base_eq.include:
            return EarthquakeConfig(include=True, a_h=0.3*base_eq.a_h, a_v=base_eq.a_v)
        else:
            return EarthquakeConfig(include=False)

    results_out = [
        _solve_case(cn, geom, mat, wl_us, wl_ds,
                    drainage, silt, backfill, req.ice_pressure, req.ice_include,
                    rock_bolt, rock_anchor, applied, incl,
                    req.rock_bolt_apply_depth_limit, req.rock_bolt_depth_limit,
                    earthquake=eq_for(cn))
        for cn, wl_us, wl_ds, incl in cases
    ]
    return SectionResponse(
        label=sec.label,
        heel_elevation=geom.heel_elevation,
        toe_elevation=geom.toe_elevation,
        base_width=round(geom.base_length_horizontal, 3),
        dam_height=round(geom.dam_top_elevation_rel, 3),
        results=results_out,
    )


@app.post("/calculate_heights", response_model=MultiHeightResponse)
def calculate_heights(req: MultiHeightRequest):
    try:
        out = [_run_section(req.base_request, s) for s in req.sections]
        return MultiHeightResponse(sections=out)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception:
        raise HTTPException(status_code=500, detail=traceback.format_exc())


class ExportHeightsRequest(BaseModel):
    # The main dam result (so it appears as a sheet too) + the section results
    main: ExportRequest
    multi: MultiHeightResponse


@app.post("/export_heights")
def export_heights(data: ExportHeightsRequest):
    """One sheet per dam height (main dam first, then each section)."""
    try:
        blocks = [_HeightBlock(
            label="Main Dam",
            heel_elevation=data.main.geometry.heel_elevation,
            toe_elevation=data.main.geometry.toe_elevation,
            base_width=data.main.geometry.base_length_horizontal,
            dam_height=data.main.geometry.dam_height,
            results=data.main.results,
        )]
        for sec in data.multi.sections:
            blocks.append(_HeightBlock(
                label=sec.label,
                heel_elevation=sec.heel_elevation,
                toe_elevation=sec.toe_elevation,
                base_width=sec.base_width,
                dam_height=sec.dam_height,
                results=sec.results,
            ))
        wb = _build_workbook(title="MULTI-HEIGHT DAM STABILITY — RESULTS",
                             blocks=blocks,
                             params=data.main.params)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(
            buf,
            media_type=("application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet"),
            headers={"Content-Disposition":
                     'attachment; filename="dam_multi_height_results.xlsx"'},
        )
    except Exception:
        raise HTTPException(status_code=500, detail=traceback.format_exc())


class DetailedCalcRequest(BaseModel):
    """Request body for the detailed calculation viewer."""
    request_payload: dict   # original LoadCaseRequest fields as dict
    results: List[CaseResult]
    geometry: GeometryInfo


@app.post("/detailed_calc")
def detailed_calc(data: DetailedCalcRequest):
    """
    Generate full detailed calculation text and return it as an HTML page
    with <pre> formatting so the user can read and print it from the browser.
    """
    try:
        p = data.request_payload

        # Reconstruct objects from the stored request payload
        coords_raw = p.get('coordinates', [])
        coords = [(c['x'], c['y']) for c in coords_raw]
        utp_raw = p.get('upstream_top_point', {})
        utp = (utp_raw.get('x', 0), utp_raw.get('y', 0))

        geom = DamGeometry(coords, utp)
        mat  = MaterialProperties(
            unit_weight_dam   = p.get('unit_weight_dam', 24.0),
            unit_weight_water = p.get('unit_weight_water', 10.0),
            friction_coeff    = p.get('friction_coeff', 0.75),
        )
        drainage = DrainageConfig(
            include              = p.get('drainage_include', False),
            distance_from_heel   = p.get('drainage_distance_from_heel', 0.0),
            reduction_factor     = p.get('drainage_reduction_factor', 0.333),
        )
        silt = SiltConfig(
            include                 = p.get('silt_include', False),
            height_us               = max(p.get('silt_elevation_us', 0) - geom.heel_elevation, 0),
            unit_weight_submerged   = p.get('silt_unit_weight_submerged', 9.0),
            phi_deg                 = p.get('silt_phi_deg', 30.0),
        )
        backfill = BackfillConfig(
            include_pressure     = p.get('backfill_pressure', False),
            include_weight       = p.get('backfill_weight', False),
            height               = p.get('backfill_height', 0.0),
            coeff_pressure       = p.get('backfill_coeff', 0.333),
            unit_weight_dry      = p.get('backfill_unit_weight_dry', 18.0),
            unit_weight_submerged= p.get('backfill_unit_weight_sub', 10.0),
        )
        ice = IcePressureConfig(
            include  = p.get('ice_include', False),
            pressure = p.get('ice_pressure', 150.0),
        )
        rock_bolt = RockBoltConfig(
            include          = p.get('rock_bolt_include', False),
            force_per_m      = p.get('rock_bolt_force_per_m', 0.0),
            cover_from_heel  = p.get('rock_bolt_cover_from_heel', 0.0),
        )
        rock_anchor = RockAnchorConfig(
            include          = p.get('rock_anchor_include', False),
            force_per_m      = p.get('rock_anchor_force_per_m', 0.0),
            cover_from_heel  = p.get('rock_anchor_cover_from_heel', 0.0),
        )
        applied = AppliedForceConfig(
            vertical_forces  = [tuple(v) for v in p.get('applied_vertical_forces', [])],
            horizontal_forces= [tuple(h) for h in p.get('applied_horizontal_forces', [])],
        )

        # Water levels per case
        wl_map = {
            'HRV':               (p.get('HRV_us', 0), p.get('HRV_ds', 0)),
            'DFV':               (p.get('DFV_us', 0), p.get('DFV_ds', 0)),
            'MFV':               (p.get('MFV_us', 0), p.get('MFV_ds', 0)),
            'DFV (no rock bolts)':(p.get('DFV_us', 0), p.get('DFV_ds', 0)),
            'HRV+EQ (X-dom)':   (p.get('HRV_us', 0), p.get('HRV_ds', 0)),
            'HRV+EQ (Y-dom)':   (p.get('HRV_us', 0), p.get('HRV_ds', 0)),
        }

        # EQ config per case
        base_ah = p.get('eq_a_h', 0.0)
        base_av = p.get('eq_a_v', 0.0)
        run_eq  = p.get('run_EQ', False)

        def eq_for(case_name):
            if case_name == 'HRV+EQ (X-dom)' and run_eq:
                return EarthquakeConfig(include=True, a_h=base_ah, a_v=0.3*base_av)
            elif case_name == 'HRV+EQ (Y-dom)' and run_eq:
                return EarthquakeConfig(include=True, a_h=0.3*base_ah, a_v=base_av)
            return EarthquakeConfig(include=False)

        # Re-run each case to get the full res dict with intermediate values
        all_text = []
        for cr in data.results:
            cn = cr.case_name
            wl_us, wl_ds = wl_map.get(cn, (p.get('HRV_us', 0), p.get('HRV_ds', 0)))
            incl_rb = (cn != 'DFV (no rock bolts)')
            eq_cfg  = eq_for(cn)
            ice_case = IcePressureConfig(
                include  = (ice.include and cn == 'HRV'),
                pressure = ice.pressure,
            )
            res = run_load_case(
                case_name=cn, geom=geom, mat=mat,
                wl_us=wl_us, wl_ds=wl_ds,
                drainage=drainage, silt=silt, backfill=backfill,
                ice=ice_case, rock_bolt=rock_bolt, rock_anchor=rock_anchor,
                applied=applied, include_rock_bolts=incl_rb,
                rb_depth_limit_apply=p.get('rock_bolt_apply_depth_limit', True),
                rb_depth_limit=p.get('rock_bolt_depth_limit', 7.0),
                earthquake=eq_cfg,
                fs_uls=p.get('fs_uls', 1.5), fs_als=p.get('fs_als', 1.1),
                res_uls=p.get('res_uls', 'middle_third'),
                res_als=p.get('res_als', 'l6'),
            )
            res['earthquake'] = eq_cfg
            text = generate_detailed_calc(
                res, geom, mat, wl_us, wl_ds,
                drainage, silt, backfill, ice_case,
                rock_bolt, rock_anchor, applied, eq_cfg)
            all_text.append(text)

        full_text = '\n\n'.join(all_text)

        # Wrap in a clean printable HTML page
        title_str = "GRAVITY DAM — DETAILED STABILITY CALCULATION"
        html = f"""<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <title>{title_str}</title>
  <style>
    body  {{ font-family: 'Courier New', Courier, monospace; font-size: 11px;
             margin: 20mm 15mm; color: #111; background: white; }}
    pre   {{ white-space: pre-wrap; word-break: break-word; line-height: 1.6; }}
    h1    {{ font-family: Arial, sans-serif; font-size: 14px; margin-bottom: 4px; }}
    p.sub {{ font-family: Arial, sans-serif; font-size: 11px; color: #555;
             margin-top: 0; margin-bottom: 16px; }}
    @media print {{
      body {{ margin: 10mm; font-size: 10px; }}
      .no-print {{ display: none; }}
    }}
  </style>
</head>
<body>
  <h1>{title_str}</h1>
  <p class="sub">Generated by Gravity Dam Stability Calculator
     &nbsp;|&nbsp; Use Ctrl+P / Cmd+P to print</p>
  <div class="no-print" style="margin-bottom:12px">
    <button onclick="window.print()"
      style="padding:6px 18px;font-size:12px;cursor:pointer;
             background:#2e6da4;color:white;border:none;border-radius:4px">
      🖨 Print
    </button>
  </div>
  <pre>{full_text}</pre>
</body>
</html>"""

        from fastapi.responses import HTMLResponse
        return HTMLResponse(content=html)

    except Exception:
        raise HTTPException(status_code=500, detail=traceback.format_exc())
